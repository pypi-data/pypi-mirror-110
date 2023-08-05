import os
import sys
import logging
import colorlog
import openpyxl
from shutil import move
from ai_object_detection import WebBrowser

class ExcelReader:
  def __init__(self, filename, downloadDir, pathPhotos, pathCaptcha = None):
    self.logger = logging.getLogger('__main__.' + __name__);
    self.current_path = os.path.abspath(os.getcwd())
    self.dataFilename = filename
    self.dataBKFilename = filename.replace('.xlsx','_backup.xlsx')
    self.pathDownload = downloadDir
    self.pathPhotos = pathPhotos
    self.pathCaptcha = pathCaptcha

    self.pathANH = os.path.join(self.current_path, 'ANH')
    self.pathPhotosDone = os.path.join(self.current_path, 'ANH_DA_XONG')
    if not os.path.exists(self.pathPhotosDone):
      os.makedirs(self.pathPhotosDone, exist_ok = True)
      self.logger.debug("Directory '%s' created" % self.pathPhotosDone)

    self.logger.info("Xu ly file: %s" % self.dataFilename)
    self.book = openpyxl.load_workbook(self.dataFilename)
    self.sheet = self.book.active

    self.header_row = 0
    for index in range(1, 4):
      ho_ten = self.sheet.cell(row=index, column=2).value
      ma_so = self.sheet.cell(row=index, column=3).value
      cmnd = self.sheet.cell(row=index, column=4).value
      tinh_tp = self.sheet.cell(row=index, column=5).value
      huyen = self.sheet.cell(row=index, column=6).value
      xa = self.sheet.cell(row=index, column=7).value
      dia_chi = self.sheet.cell(row=index, column=8).value
      so_dt = self.sheet.cell(row=index, column=9).value
      row_done = self.sheet.cell(row=index, column=10).value

      if ho_ten == "hoTen" and cmnd == "soCmnd":
        self.header_row = index
        break
    self.book.close()

    if self.header_row == 0:
      self.logger.error("File excel khong dung dinh dang: maDvi	hoTen	soBhxh	soCmnd	maTinhLh	maHuyenLh	maXaLh	dia chi	soDienThoai")
      os._exit(1)
    """
    try:
      if (row_done and int(row_done)) > 1:
        self.start_row = int(row_done)
    except:
      self.start_row =  self.header_row + 1
      self.logger.info("Khong co thong tin dong cuoi cung da lam")
    """
    self.start_row =  self.header_row + 1

  def archive_cmnd_files(self,filename):
    srcFileAnhCmndMatTruoc = os.path.join(self.pathPhotos,filename)
    dstFileAnhCmndMatTruoc = os.path.join(self.pathPhotosDone,filename)

    srcFileAnhCmndMatSau = os.path.join(self.pathPhotos, filename.replace('MT','MS'))
    dstFileAnhCmndMatSau = os.path.join(self.pathPhotosDone, filename.replace('MT','MS'))

    srcFileAnhCmndCD = os.path.join(self.pathPhotos,filename.replace('MT','CD'))
    dstFileAnhCmndCD = os.path.join(self.pathPhotosDone,filename.replace('MT','CD'))

    if (os.path.exists(srcFileAnhCmndMatTruoc)):
      self.logger.info("hoan thanh file: %s" % os.path.basename(srcFileAnhCmndMatTruoc))
      move(srcFileAnhCmndMatTruoc, dstFileAnhCmndMatTruoc)
    if (os.path.exists(srcFileAnhCmndMatSau)):
      self.logger.info("hoan thanh file: %s" % os.path.basename(srcFileAnhCmndMatSau))
      move(srcFileAnhCmndMatSau, dstFileAnhCmndMatSau)
    if (os.path.exists(srcFileAnhCmndCD)):
      self.logger.info("hoan thanh file: %s" % os.path.basename(srcFileAnhCmndCD))
      move(srcFileAnhCmndCD, dstFileAnhCmndCD)

  def renameFileName(self, column_filename=1, column_cmnd=4):
    self.book = openpyxl.load_workbook(self.dataFilename)
    self.sheet = self.book.active
    for index in range(self.header_row + 1, 2000):
      try:
        old_filename = self.sheet.cell(row=index, column=column_filename).value
        cmnd = self.sheet.cell(row=index, column=column_cmnd).value
        if not cmnd:
          self.logger.info("[%i] Ket thuc file" % (index))
          break
        if not old_filename:
          continue
        self.logger.info("'%s' CD,MT,MS -> '%s' CD,MT,MS" % (old_filename,cmnd))

        list_ext = ['.pdf', '.jpg', '.png', '.jpeg']
        list_prefix = ['', '0', '00']
        list_suffix = ['','CD', 'MT', 'MS', ' ',' CD', ' MT', ' MS']
        is_found = False
        for extension in list_ext:
          for prefix in list_prefix:
            for suffix in list_suffix:
              old_filepath = os.path.join(self.pathANH,'%s%s%s%s' % (prefix,old_filename,suffix,extension))
              new_filepath = os.path.join(self.pathANH,'%s%s%s%s' % (prefix,cmnd,suffix,extension))
              new_filepath = new_filepath.replace(' ', '')
              if (os.path.exists(old_filepath)):
                is_found = True
                if (os.path.exists(new_filepath)):
                  self.logger.warn("New_File da ton tai '%s'" % os.path.basename(new_filepath))
                else:
                  self.logger.info("'%s' -> '%s'" % (os.path.basename(old_filepath),os.path.basename(new_filepath)))
                  move(old_filepath, new_filepath)
            if is_found:
              break
          if is_found:
            break
        if not is_found:
          self.logger.debug("Old_File khong ton tai '%s...'" % old_filename)

      except AttributeError as err:
        self.logger.info ("Attribute Error: {0}".format(err))
        break
      except NameError as err:
        self.logger.info ("Name Error: {0}".format(err))
        break
      except TypeError as err:
        self.logger.info ("Type Error: {0}".format(err))
        break
      except FileNotFoundError as err:
        self.logger.info ("Type Error: {0}".format(err))
        break
      except KeyboardInterrupt as err:
        self.logger.info ("KeyboardInterrupt: {0}".format(err))
        break
      except:
        self.logger.error("Unexpected error: %s" % sys.exc_info()[0])
    self.book.close()

  def analyze(self):
    # self.analyzeFolder()
    self.analyzeExcel()

  def analyzeFolder(self):
    self.logger.info("Xu ly thu muc: %s" % self.pathPhotos)

    list_of_dir = os.listdir(self.pathPhotos)
    if len(list_of_dir) == 0:
      self.logger.info("Chua co tap tin CD,MT,MS trong %s" % self.pathPhotos)
      return None

    self.webBrowser = WebBrowser(self.pathDownload)
    self.webBrowser.getConfig()
    self.webBrowser.startDriver()

    for filename in list_of_dir:
      self.logger.info("Xu ly file: %s" % filename)

  def analyzeExcel(self):
    self.webBrowser = WebBrowser(self.pathDownload,self.pathCaptcha)
    self.webBrowser.getConfig()
    self.webBrowser.startDriver()

    self.book = openpyxl.load_workbook(self.dataFilename)
    self.sheet = self.book.active
    self.book.save(self.dataBKFilename)
    is_excel_opened = True
    for index in range(self.start_row, 2000):
      try:
        if not is_excel_opened:
          self.book = openpyxl.load_workbook(self.dataFilename)
          self.sheet = self.book.active
          is_excel_opened = True

        ho_ten = self.sheet.cell(row=index, column=2).value
        ma_so = self.sheet.cell(row=index, column=3).value
        cmnd = self.sheet.cell(row=index, column=4).value
        tinh_tp = self.sheet.cell(row=index, column=5).value
        huyen = self.sheet.cell(row=index, column=6).value
        xa = self.sheet.cell(row=index, column=7).value
        dia_chi = self.sheet.cell(row=index, column=8).value
        so_dt = self.sheet.cell(row=index, column=9).value

        if not ma_so:
          self.logger.info("[%i] Ket thuc file" % (index))
          break

        cmnd_str = str(cmnd)
        ma_so_str = str(ma_so)
        tinh_tp_str = str(tinh_tp)
        huyen_str = str(huyen)
        xa_str = str(xa)
        if not dia_chi:
          dia_chi = "."

        filename = cmnd_str + "MT.jpg"

        if self.sheet.cell(row=index, column=10).value == "xong":
          self.logger.info("[%i] '%s' Da duoc nhap" % (index,cmnd_str))
          self.archive_cmnd_files(filename)
          continue

        if self.sheet.cell(row=index, column=10).value == "error":
          self.logger.warning("[%i] Bo qua '%s' do trang thai 'error'" % (index,cmnd_str))
          continue

        if not cmnd or not ho_ten or not so_dt:
          self.logger.error("[%i] Du lieu loi" % (index))
          self.sheet.cell(row=index, column=10).value = "error"
          self.sheet.cell(row=index, column=11).value = "Du lieu loi, khong duoc cung cap day du"
          self.sheet.cell(row=self.header_row, column=10).value = index
          self.book.save(self.dataFilename)
          self.book.close()
          is_excel_opened = False
          continue

        self.logger.info(("[%i] %s : %s : %s : %s : %s : %s : %s : %s"% (index,ho_ten,ma_so_str,cmnd_str,tinh_tp_str,huyen_str,xa_str,dia_chi,so_dt)).encode('cp1250','replace'))
        print("[%i] %s : %s : %s : %s : %s : %s : %s : %s"% (index,ho_ten,ma_so_str,cmnd_str,tinh_tp_str,huyen_str,xa_str,dia_chi,so_dt))

        file_cd = os.path.join(self.pathPhotos, filename.replace('MT','CD'))
        file_mt = os.path.join(self.pathPhotos, filename)
        file_ms = os.path.join(self.pathPhotos, filename.replace('MT','MS'))

        """
        if len(cmnd_str) != 12:
          if not os.path.exists(file_mt) or not os.path.exists(file_ms):
            self.logger.error("[%i] Du lieu loi, theu anh MT va MS" % (index))
            self.sheet.cell(row=index, column=10).value = "error"
            self.sheet.cell(row=self.header_row, column=10).value = index
            self.sheet.cell(row=index, column=11).value = "Theu anh %sCD.jpg, %sMT.jpg va %sMS.jpg" % (cmnd_str,cmnd_str,cmnd_str)
            self.book.save(self.dataFilename)
            self.book.close()
            is_excel_opened = False
            continue
        """

        self.webBrowser.launchUrl('https://dichvucong.baohiemxahoi.gov.vn/#/dang-ky?loaidoituong=0')
        self.webBrowser.fillForm(file_cd, file_mt, file_ms, ho_ten, ma_so_str, cmnd_str, tinh_tp_str, huyen_str, xa_str, dia_chi, so_dt)
        status,msg,is_top = self.webBrowser.waitCaptcha(file_cd, file_mt, file_ms, ho_ten, ma_so_str, cmnd_str, tinh_tp_str, huyen_str, xa_str, dia_chi, so_dt)

        self.sheet.cell(row=index, column=10).value = status
        self.sheet.cell(row=index, column=11).value = msg
        self.sheet.cell(row=self.header_row, column=10).value = index
        self.book.save(self.dataFilename)
        self.book.close()
        is_excel_opened = False
        if status == "xong":
          self.archive_cmnd_files(filename)
        if is_top:
          break
      except AttributeError as err:
        self.logger.info ("Attribute Error: {0}".format(err))
        break
      except NameError as err:
        self.logger.info ("Name Error: {0}".format(err))
        break
      except TypeError as err:
        self.logger.info ("Type Error: {0}".format(err))
        break
      except FileNotFoundError as err:
        self.logger.info ("Type Error: {0}".format(err))
        break
      except KeyboardInterrupt as err:
        self.logger.info ("KeyboardInterrupt: {0}".format(err))
        break
      except:
        self.logger.error("Unexpected error: %s" % sys.exc_info()[0])
        break

    if is_excel_opened:
      self.logger.info ("Closing excel file")
      self.book.close()
      self.logger.info ("Closed excel file")
    self.webBrowser.quitDriver()
