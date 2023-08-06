import random
import sys
from typing import List

import openpyxl
import pandas as pd

from .connector import Connector


class Excel(Connector):
    """
    Read data from a Excel file.
    """
    TMP_SINK = False
    MAX_SIZE = 2000000
    MAX_ROWS = 200
    fields = {
        "filename": {"name": "Filename", "type": "file", "input": "file", "help": "The filename of the CSV file.",
                     "required": True},
        "sheet_name": {"name": "Sheet name", "type": "string", "input": "text", "required": True, "default": ",",
                       "help": "The name or number of the sheet."},
        "has_header": {"name": "Has header", "type": "boolean", "input": "switch", "required": True, "default": False,
                       "help": "Does the file have a header containing the column names?."},
    }

    def __init__(self, arguments):
        """Initialize the data source with the given parameters.

        Arguments:
            arguments {dict} -- The arguments
        """
        super().__init__(arguments)
        self.filename = arguments["filename"]
        self.sheet_name = arguments["sheet_name"]
        self.has_header = arguments["has_header"]

    def _get_row_values(self, row: tuple) -> list:
        return [cell.value for cell in row]

    def __call__(self, sample: bool = False, sampling_technique: str = "top", column_types: bool = False) -> List[dict]:
        """This class is called when the data needs to be loaded.

        Arguments:
            :type sample: boolean: Whether to take a sample or not
            :type sampling_technique: str: Which sampling technique to use (top, stratisfied, random)

        Returns:
            dict -- The row, including the extra output column
        """

        if sample:
            wb = openpyxl.load_workbook(self.filename, read_only=True)
            if self.sheet_name not in wb.sheetnames:
                try:
                    sheet_idx = int(self.sheet_name)
                    if sheet_idx < len(wb.sheetnames):
                        self.sheet_name = wb.sheetnames[sheet_idx]
                    else:
                        raise KeyError("The sheet doesn't exist")
                except ValueError as e:
                    raise KeyError("The sheet doesn't exist")

            sheet = wb[self.sheet_name]

            # TODO: Make this big data proof (chucking, sampling before loading, etc.)

            row_sizes = []
            for row in sheet.rows:
                # We'll test the first 25 rows to determine average row size
                row_sizes.append(sys.getsizeof(self._get_row_values(row)))

                # We want to check at least 25 rows, at most 250 and ideally 1%
                if len(row_sizes) > max(min(sheet.max_row * 0.01, 250), 25):
                    break

            sample_size = min(self.MAX_ROWS, round(self.MAX_SIZE / (sum(row_sizes) / len(row_sizes))))
            column_names, data = [], []

            i = 0
            if sampling_technique == "top":
                # We'll just take the top rows
                for row in sheet.rows:
                    if i == 0 and self.has_header:
                        column_names = self._get_row_values(row)
                    elif i <= sample_size:
                        data.append(self._get_row_values(row))
                    else:
                        break
                    i += 1
            elif sampling_technique == "stratified":
                # We'll take every xth row
                stratified = round(sheet.max_row / sample_size)
                for row in sheet.rows:
                    if i == 0 and self.has_header:
                        column_names = self._get_row_values(row)
                    elif i % stratified == 0:
                        data.append(self._get_row_values(row))
                    i += 1
            else:
                # We're doing random sampling ...
                rows_to_take = random.sample(range(1 if self.has_header else 0, sheet.max_row), sample_size)
                rows_to_take = sorted(rows_to_take)
                for row in sheet.rows:
                    if i == 0 and self.has_header:
                        column_names = self._get_row_values(row)
                    elif i == rows_to_take[0]:
                        data.append(self._get_row_values(row))
                        rows_to_take.pop(0)
                    if len(rows_to_take) == 0:
                        break
                    i += 1

            df = pd.DataFrame(data, columns=column_names)
            return df.to_dict(orient="records")
        else:
            # TODO: To be implemented properly!
            # raise NotImplementedError()
            df = pd.read_excel(self.filename, sheet_name=self.sheet_name, header=0 if self.has_header else None,
                               engine="openpyxl")
            return df.to_dict(orient="records")
